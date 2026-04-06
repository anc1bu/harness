"""Harness backend — auth + reference/transactional data store."""
from flask import Flask, request, jsonify, send_from_directory, session, redirect, Response, stream_with_context
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
from datetime import datetime, timedelta
from functools import wraps
import sqlite3, json, os, re, secrets

# Map SAP SE16N description headers for DD03L to their technical field names
_DD03L_HEADER_MAP = {
    'Table Name': 'TABNAME', 'Field Name': 'FIELDNAME',
    'Activation State': 'AS4LOCAL', 'Version': 'AS4VERS',
    'Table position': 'POSITION', 'Table position.1': 'DBPOSITION',
    'Key field': 'KEYFLAG', 'Required Field': 'MANDATORY',
    'Data element': 'ROLLNAME', 'Check table': 'CHECKTABLE',
    'Admin. field': 'ADMINFIELD', 'ABAP type': 'INTTYPE',
    'Internal Length': 'INTLEN', 'Reference table': 'REFTABLE',
    'Name of include': 'PRECFIELD', 'Ref. field': 'REFFIELD',
    'Check module': 'CONROUT', 'Force NOT NULL': 'NOTNULL',
    'Data Type': 'DATATYPE', 'No. of Characters': 'LENG',
    'Decimal Places': 'DECIMALS', 'Domain name': 'DOMNAME',
    'Origin': 'SHLPORIGIN', 'Table': 'TABLETYPE', 'Depth': 'DEPTH',
    'Component Type': 'COMPTYPE', 'Type of Object Referenced': 'REFTYPE',
    'Text Lang.': 'LANGUFLAG', 'Anonymous': 'ANONYMOUS',
    'Output': 'OUTPUTSTYLE', 'SRS Identifier': 'SRS_ID',
}

# All columns stored in _dd03l (fixed schema)
_DD03L_ALL_COLS = [
    'TABNAME', 'FIELDNAME', 'ROLLNAME', 'CHECKTABLE', 'KEYFLAG', 'MANDATORY',
    'DATATYPE', 'LENG', 'DECIMALS', 'INTTYPE', 'INTLEN', 'POSITION', 'DBPOSITION',
    'AS4LOCAL', 'AS4VERS', 'ADMINFIELD', 'REFTABLE', 'PRECFIELD', 'REFFIELD',
    'CONROUT', 'NOTNULL', 'DOMNAME', 'SHLPORIGIN', 'TABLETYPE', 'DEPTH',
    'COMPTYPE', 'REFTYPE', 'LANGUFLAG', 'ANONYMOUS', 'OUTPUTSTYLE', 'SRS_ID',
]

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.environ.get('HARNESS_DATA', os.path.join(HERE, 'data'))
os.makedirs(DATA, exist_ok=True)

app = Flask(__name__, static_folder=HERE, static_url_path='')
app.secret_key = os.environ.get('HARNESS_SECRET', secrets.token_hex(32))
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200 MB
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=14)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Dev bypass: when HARNESS_NO_AUTH=1, skip all auth checks.
NO_AUTH = os.environ.get('HARNESS_NO_AUTH', '').strip() in ('1', 'true', 'yes')
DEV_EMAIL = 'dev@harness.local'

# ───────────── AUTH DB (users.db) ─────────────
def db():
    c = sqlite3.connect(os.path.join(DATA, 'users.db'))
    c.row_factory = sqlite3.Row
    return c

def init_db():
    c = db()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY,
        email TEXT UNIQUE NOT NULL,
        pwhash TEXT NOT NULL,
        created_at TEXT NOT NULL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS projects (
        id         INTEGER PRIMARY KEY,
        name       TEXT UNIQUE NOT NULL COLLATE NOCASE,
        created_at TEXT NOT NULL)''')
    c.execute('INSERT OR IGNORE INTO projects (name, created_at) VALUES (?,?)',
              ('Dev', datetime.utcnow().isoformat() + 'Z'))
    c.commit(); c.close()
init_db()

# ───────────── HARNESS DB (harness.db) ─────────────
HARNESS_DB = os.path.join(DATA, 'harness.db')

def harness_db():
    c = sqlite3.connect(HARNESS_DB)
    c.row_factory = sqlite3.Row
    c.execute('PRAGMA journal_mode=WAL')
    return c

def init_harness_db():
    c = harness_db()
    cols_ddl = ', '.join(
        f'"{col}" TEXT' if col not in ('TABNAME', 'FIELDNAME') else f'"{col}" TEXT NOT NULL'
        for col in _DD03L_ALL_COLS
    )
    c.executescript(f'''
        CREATE TABLE IF NOT EXISTS _dd03l (
            {cols_ddl},
            PRIMARY KEY ("TABNAME", "FIELDNAME")
        );
        CREATE TABLE IF NOT EXISTS _dd03l_meta (k TEXT PRIMARY KEY, v TEXT);

        CREATE TABLE IF NOT EXISTS dd04t (
            rollname TEXT PRIMARY KEY,
            SCRTEXT_M TEXT, SCRTEXT_L TEXT, SCRTEXT_S TEXT,
            DDTEXT TEXT, REPTEXT TEXT
        );
        CREATE TABLE IF NOT EXISTS dd04t_meta (k TEXT PRIMARY KEY, v TEXT);

        CREATE TABLE IF NOT EXISTS _dd08l (
            CHECKTABLE TEXT, TABNAME TEXT, FRKART TEXT, FIELDNAME TEXT,
            _extra TEXT
        );
        CREATE TABLE IF NOT EXISTS _dd08l_meta (k TEXT PRIMARY KEY, v TEXT);

        CREATE TABLE IF NOT EXISTS _table_meta (
            tablename   TEXT PRIMARY KEY,
            system      TEXT,
            client      TEXT,
            date        TEXT,
            project     TEXT,
            filename    TEXT,
            uploadedAt  TEXT,
            columns     TEXT,
            enriched_columns TEXT
        );
    ''')
    c.commit(); c.close()
init_harness_db()

def login_required(f):
    @wraps(f)
    def w(*a, **kw):
        if NO_AUTH:
            session['project'] = 'Dev'
            return f(*a, **kw)
        if 'uid' not in session:
            return jsonify(error='auth_required'), 401
        return f(*a, **kw)
    return w

# ───────────── AUTH ─────────────
@app.post('/api/auth/signup')
def signup():
    d = request.get_json(silent=True) or {}
    email = (d.get('email') or '').strip().lower()
    pw = d.get('password') or ''
    if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return jsonify(error='invalid_email'), 400
    if len(pw) < 6:
        return jsonify(error='password_too_short'), 400
    c = db()
    try:
        c.execute('INSERT INTO users (email,pwhash,created_at) VALUES (?,?,?)',
                  (email, generate_password_hash(pw, method='pbkdf2:sha256'), datetime.utcnow().isoformat()))
        c.commit()
    except sqlite3.IntegrityError:
        c.close(); return jsonify(error='email_exists'), 409
    row = c.execute('SELECT id FROM users WHERE email=?', (email,)).fetchone()
    c.close()
    session.permanent = True
    session['uid'] = row['id']; session['email'] = email
    return jsonify(ok=True, email=email)

@app.post('/api/auth/login')
def login():
    d = request.get_json(silent=True) or {}
    email = (d.get('email') or '').strip().lower()
    pw = d.get('password') or ''
    c = db()
    row = c.execute('SELECT id,pwhash FROM users WHERE email=?', (email,)).fetchone()
    c.close()
    if not row or not check_password_hash(row['pwhash'], pw):
        return jsonify(error='bad_credentials'), 401
    session.permanent = True
    session['uid'] = row['id']; session['email'] = email
    return jsonify(ok=True, email=email)

@app.post('/api/auth/logout')
def logout():
    session.clear()
    return jsonify(ok=True)

@app.get('/api/projects')
def list_projects():
    if not NO_AUTH and 'uid' not in session:
        return jsonify(error='auth_required'), 401
    c = db()
    rows = c.execute('SELECT name FROM projects ORDER BY name COLLATE NOCASE').fetchall()
    c.close()
    return jsonify(projects=[r['name'] for r in rows], current=session.get('project'))

@app.post('/api/auth/set-project')
@login_required
def set_project():
    d = request.get_json(silent=True) or {}
    name = (d.get('project') or '').strip()[:64]
    if not name:
        return jsonify(error='missing_project', hint='Project name is required'), 400
    create = bool(d.get('create', False))
    c = db()
    if create:
        c.execute('INSERT OR IGNORE INTO projects (name, created_at) VALUES (?,?)',
                  (name, datetime.utcnow().isoformat() + 'Z'))
        c.commit()
    else:
        row = c.execute('SELECT name FROM projects WHERE name=? COLLATE NOCASE', (name,)).fetchone()
        if not row:
            c.close()
            return jsonify(error='project_not_found', hint=f'Project "{name}" does not exist'), 404
        name = row['name']
    c.close()
    session['project'] = name
    return jsonify(ok=True, project=name)

@app.get('/api/auth/me')
def me():
    if NO_AUTH:
        return jsonify(authed=True, email=DEV_EMAIL, dev=True, project='Dev')
    if 'uid' not in session:
        return jsonify(authed=False)
    return jsonify(authed=True, email=session.get('email'), project=session.get('project'))

# ───────────── DD03L HELPERS ─────────────
def _read_dd03l():
    """Return {rows, tabnames, uploadedAt, filename} from harness.db or None."""
    c = harness_db()
    try:
        meta = dict(c.execute('SELECT k,v FROM _dd03l_meta').fetchall())
    except sqlite3.OperationalError:
        c.close(); return None
    if not meta.get('uploadedAt'):
        c.close(); return None
    col_sql = ', '.join(f'"{col}"' for col in _DD03L_ALL_COLS)
    raw = c.execute(f'SELECT {col_sql} FROM _dd03l').fetchall()
    c.close()
    rows = [dict(zip(_DD03L_ALL_COLS, r)) for r in raw]
    tabnames = sorted({r['TABNAME'] for r in rows if r.get('TABNAME')})
    return {'rows': rows, 'tabnames': tabnames,
            'uploadedAt': meta.get('uploadedAt'), 'filename': meta.get('filename')}

def _write_dd03l(slim, filename):
    """Merge slim rows into _dd03l, replacing rows for tables present in slim."""
    uploaded_tabnames = {r['TABNAME'] for r in slim}
    ph = ','.join('?' * len(uploaded_tabnames))
    col_list = ', '.join(f'"{col}"' for col in _DD03L_ALL_COLS)
    val_marks = ', '.join('?' * len(_DD03L_ALL_COLS))
    c = harness_db()
    c.execute('BEGIN')
    c.execute(f'DELETE FROM _dd03l WHERE "TABNAME" IN ({ph})', list(uploaded_tabnames))
    c.executemany(
        f'INSERT OR REPLACE INTO _dd03l ({col_list}) VALUES ({val_marks})',
        [tuple(r.get(col, '') for col in _DD03L_ALL_COLS) for r in slim]
    )
    c.execute('INSERT OR REPLACE INTO _dd03l_meta VALUES (?,?)',
              ('uploadedAt', datetime.utcnow().isoformat() + 'Z'))
    c.execute('INSERT OR REPLACE INTO _dd03l_meta VALUES (?,?)', ('filename', filename))
    c.commit(); c.close()

# ───────────── STATUS ─────────────
@app.get('/api/status')
@login_required
def status():
    dd03l = _read_dd03l()
    dd03l_info = None
    if dd03l:
        dd03l_info = {'loaded': True, 'rows': len(dd03l.get('rows', [])),
                      'uploadedAt': dd03l.get('uploadedAt'), 'filename': dd03l.get('filename'),
                      'tabnames': dd03l.get('tabnames', [])}
    dd04t_info = _dd04t_info()
    dd08l_info = _dd08l_info()
    current_project = 'Dev' if NO_AUTH else session.get('project')
    c = harness_db()
    meta_rows = c.execute(
        'SELECT tablename, columns, filename, uploadedAt, system, client, date '
        'FROM _table_meta WHERE project=? ORDER BY tablename',
        (current_project,)
    ).fetchall()
    tables = []
    for m in meta_rows:
        cols = json.loads(m['columns'] or '[]')
        try:
            row_count = c.execute(f'SELECT COUNT(*) FROM "{m["tablename"]}"').fetchone()[0]
        except sqlite3.OperationalError:
            row_count = 0
        tables.append({'name': m['tablename'], 'rows': row_count, 'columns': len(cols),
                       'filename': m['filename'], 'uploadedAt': m['uploadedAt'],
                       'system': m['system'], 'client': m['client'], 'date': m['date']})
    c.close()
    return jsonify(dd03l=dd03l_info, dd04t=dd04t_info, dd08l=dd08l_info, tables=tables)

# ───────────── UPLOADS ─────────────
def _parse_wb(fileobj):
    wb = load_workbook(fileobj, read_only=True, data_only=True)
    out = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = None
        for r in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(x).strip() if x is not None else '' for x in r]
                continue
            if all(v is None or v == '' for v in r): continue
            out.append({h: ('' if v is None else v) for h, v in zip(headers, r)})
    return out, headers or []

@app.post('/api/upload/dd03l')
@login_required
def up_dd03l():
    f = request.files.get('file')
    if not f: return jsonify(error='no_file'), 400
    rows, _ = _parse_wb(f)
    rows = [{_DD03L_HEADER_MAP.get(k, k): v for k, v in r.items()} for r in rows]
    slim = []
    for r in rows:
        tn = str(r.get('TABNAME', '')).strip()
        fn = str(r.get('FIELDNAME', '')).strip()
        if not tn or not fn: continue
        slim.append({k: (str(v).strip() if v is not None else '') for k, v in r.items()})
    if not slim:
        return jsonify(error='no_valid_rows'), 400
    _write_dd03l(slim, f.filename)
    _reenrich_all()
    c = harness_db()
    total = c.execute('SELECT COUNT(*) FROM _dd03l').fetchone()[0]
    tab_count = c.execute('SELECT COUNT(DISTINCT "TABNAME") FROM _dd03l').fetchone()[0]
    c.close()
    uploaded_tabnames = {r['TABNAME'] for r in slim}
    return jsonify(ok=True, rows=total, tabnames=tab_count, merged=list(uploaded_tabnames))

@app.post('/api/upload/dd04t')
@login_required
def up_dd04t():
    f = request.files.get('file')
    if not f: return jsonify(error='no_file'), 400
    tmp_xlsx = os.path.join(DATA, f'_dd04t_upload_{secrets.token_hex(4)}.xlsx')
    f.save(tmp_xlsx)
    original_name = f.filename

    def gen():
        c = None
        try:
            c = harness_db()
            c.execute('PRAGMA synchronous=NORMAL')
            c.execute('BEGIN EXCLUSIVE')
            c.execute('DELETE FROM dd04t')
            c.execute('DELETE FROM dd04t_meta')
            wb = load_workbook(tmp_xlsx, read_only=True, data_only=True)
            eng = {'EN', 'E', 'en', 'e'}
            seen = set(); n = 0; BATCH = 5000; batch = []
            yield json.dumps({'event': 'start'}) + '\n'
            for sn in wb.sheetnames:
                ws = wb[sn]
                headers = None; idx = {}
                for r in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(x).strip() if x is not None else '' for x in r]
                        idx = {h: i for i, h in enumerate(headers)}
                        continue
                    if 'ROLLNAME' not in idx or 'DDLANGUAGE' not in idx: break
                    lang_v = r[idx['DDLANGUAGE']]
                    if lang_v is None or str(lang_v).strip() not in eng: continue
                    rn = r[idx['ROLLNAME']]
                    if rn is None: continue
                    rn = str(rn).strip()
                    if not rn or rn in seen: continue
                    seen.add(rn)
                    def cell(key):
                        i = idx.get(key, -1)
                        if i < 0: return ''
                        v = r[i]
                        return '' if v is None else str(v).strip()
                    batch.append((rn, cell('SCRTEXT_M'), cell('SCRTEXT_L'),
                                  cell('SCRTEXT_S'), cell('DDTEXT'), cell('REPTEXT')))
                    n += 1
                    if len(batch) >= BATCH:
                        c.executemany('INSERT OR IGNORE INTO dd04t VALUES (?,?,?,?,?,?)', batch)
                        batch.clear()
                        yield json.dumps({'event': 'progress', 'rows': n}) + '\n'
            if batch:
                c.executemany('INSERT OR IGNORE INTO dd04t VALUES (?,?,?,?,?,?)', batch)
            c.execute('INSERT OR REPLACE INTO dd04t_meta VALUES (?,?)',
                      ('uploadedAt', datetime.utcnow().isoformat() + 'Z'))
            c.execute('INSERT OR REPLACE INTO dd04t_meta VALUES (?,?)', ('filename', original_name))
            c.execute('INSERT OR REPLACE INTO dd04t_meta VALUES (?,?)', ('rows', str(n)))
            c.commit(); c.close(); c = None
            _reenrich_all()
            yield json.dumps({'event': 'done', 'ok': True, 'rows': n}) + '\n'
        except Exception as e:
            if c:
                try: c.rollback(); c.close()
                except Exception: pass
            yield json.dumps({'event': 'error', 'error': str(e)}) + '\n'
        finally:
            try: os.remove(tmp_xlsx)
            except OSError: pass

    return Response(stream_with_context(gen()), mimetype='application/x-ndjson',
                    headers={'X-Accel-Buffering': 'no', 'Cache-Control': 'no-cache'})

def _dd04t_info():
    c = harness_db()
    try:
        m = dict(c.execute('SELECT k,v FROM dd04t_meta').fetchall())
    except sqlite3.OperationalError:
        c.close(); return None
    c.close()
    rows = int(m.get('rows', 0))
    if rows == 0: return None
    return {'loaded': True, 'rows': rows,
            'uploadedAt': m.get('uploadedAt'), 'filename': m.get('filename')}

def _dd04t_lookup(rollnames, textfield):
    """Return {rollname: text} for given set of rollnames."""
    if not rollnames: return {}
    if textfield not in ('SCRTEXT_M', 'SCRTEXT_L', 'SCRTEXT_S', 'DDTEXT', 'REPTEXT'):
        textfield = 'SCRTEXT_M'
    c = harness_db()
    out = {}
    rollnames = list(rollnames)
    for i in range(0, len(rollnames), 500):
        chunk = rollnames[i:i+500]
        q = f'SELECT rollname,"{textfield}" FROM dd04t WHERE rollname IN ({",".join("?"*len(chunk))})'
        for rn, txt in c.execute(q, chunk):
            if txt: out[rn] = txt
    c.close()
    return out

# ───────────── ENRICHMENT HELPERS ─────────────
def _enrich_columns(table, columns, textfield='SCRTEXT_M'):
    """Return {fieldname: 'FIELDNAME - description'} for all resolvable columns."""
    c = harness_db()
    rows = c.execute(
        'SELECT "FIELDNAME", "ROLLNAME" FROM _dd03l WHERE "TABNAME"=?', (table,)
    ).fetchall()
    c.close()
    f2r = {r[0]: r[1] for r in rows if r[0] and r[1]}
    needed_rolls = set(f2r.values())
    txts = _dd04t_lookup(needed_rolls, textfield)
    enriched = {}
    for col in columns:
        roll = f2r.get(col, '')
        txt = txts.get(roll, '') if roll else ''
        if txt:
            enriched[col] = f'{col} - {txt}'
    return enriched

def _reenrich_all():
    """Re-run enrichment for all stored transactional tables."""
    c = harness_db()
    meta_rows = c.execute(
        'SELECT tablename, columns, enriched_columns FROM _table_meta'
    ).fetchall()
    c.close()
    for m in meta_rows:
        table = m['tablename']
        cols  = json.loads(m['columns'] or '[]')
        if not table or not cols: continue
        try:
            new_enriched = _enrich_columns(table, cols)
            existing = json.loads(m['enriched_columns'] or '{}') or {}
            existing.update(new_enriched)
            c2 = harness_db()
            c2.execute('UPDATE _table_meta SET enriched_columns=? WHERE tablename=?',
                       (json.dumps(existing), table))
            c2.commit(); c2.close()
        except Exception:
            pass

# ───────────── DD08L ─────────────
def _dd08l_info():
    c = harness_db()
    try:
        m = dict(c.execute('SELECT k,v FROM _dd08l_meta').fetchall())
    except sqlite3.OperationalError:
        c.close(); return None
    c.close()
    rows = int(m.get('rows', 0))
    if rows == 0: return None
    return {'loaded': True, 'rows': rows,
            'uploadedAt': m.get('uploadedAt'), 'filename': m.get('filename')}

def _dd08l_lookup(checktable):
    """Return the text-table TABNAME for a given check table (FRKART=TEXT), or None."""
    c = harness_db()
    row = c.execute(
        'SELECT TABNAME FROM _dd08l WHERE CHECKTABLE=? AND FRKART=? LIMIT 1',
        (checktable, 'TEXT')
    ).fetchone()
    c.close()
    if not row: return None
    return row[0].strip() or None

@app.post('/api/upload/dd08l')
@login_required
def up_dd08l():
    f = request.files.get('file')
    if not f: return jsonify(error='no_file'), 400
    rows, _ = _parse_wb(f)
    _DD08L_FIXED = {'CHECKTABLE', 'TABNAME', 'FRKART', 'FIELDNAME'}
    batch = []
    for r in rows:
        clean = {k: (str(v).strip() if v is not None else '') for k, v in r.items()}
        extra = {k: v for k, v in clean.items() if k not in _DD08L_FIXED}
        batch.append((clean.get('CHECKTABLE', ''), clean.get('TABNAME', ''),
                      clean.get('FRKART', ''), clean.get('FIELDNAME', ''),
                      json.dumps(extra) if extra else None))
    c = harness_db()
    c.execute('BEGIN')
    c.execute('DELETE FROM _dd08l')
    c.execute('DELETE FROM _dd08l_meta')
    c.executemany(
        'INSERT INTO _dd08l (CHECKTABLE, TABNAME, FRKART, FIELDNAME, _extra) VALUES (?,?,?,?,?)',
        batch)
    c.execute('INSERT OR REPLACE INTO _dd08l_meta VALUES (?,?)',
              ('uploadedAt', datetime.utcnow().isoformat() + 'Z'))
    c.execute('INSERT OR REPLACE INTO _dd08l_meta VALUES (?,?)', ('filename', f.filename))
    c.execute('INSERT OR REPLACE INTO _dd08l_meta VALUES (?,?)', ('rows', str(len(batch))))
    c.commit(); c.close()
    return jsonify(ok=True, rows=len(batch))

FNAME_RE = re.compile(r'^([A-Z0-9]+)_([A-Z0-9]+)_(\d+)_(\d{8})\.xlsx?$', re.I)

@app.post('/api/upload/trans')
@login_required
def up_trans():
    current_project = 'Dev' if NO_AUTH else session.get('project')
    if not current_project:
        return jsonify(error='no_project',
                       hint='Select or create a project before uploading tables.'), 400
    f = request.files.get('file')
    if not f: return jsonify(error='no_file'), 400
    m = FNAME_RE.match(f.filename)
    if not m:
        return jsonify(error='bad_filename',
                       hint='Expected TABLE_SYSTEM_CLIENT_YYYYMMDD.xlsx'), 400
    table, system, client, date = (m.group(1).upper(), m.group(2).upper(),
                                   m.group(3), m.group(4))
    try: datetime.strptime(date, '%Y%m%d')
    except ValueError: return jsonify(error='bad_date'), 400
    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = None; rows = []
    for r in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(x).strip() if x is not None else '' for x in r]
            continue
        if all(v is None or v == '' for v in r): continue
        rows.append({h: ('' if v is None else v) for h, v in zip(headers, r)})

    # Column validation
    dd03l = _read_dd03l()
    if not dd03l:
        return jsonify(error='dd03l_required',
                       hint='DD03L is not loaded. Upload DD03L before uploading any table.'), 400
    tbl_fields = {r['FIELDNAME'] for r in dd03l.get('rows', [])
                  if r.get('TABNAME') == table and r.get('FIELDNAME')}
    if not tbl_fields:
        is_dd03l_self = (
            table == 'DD03L' and bool(rows) and
            all(str(r.get('TABNAME', '')).strip().upper() == 'DD03L' for r in rows)
        )
        if not is_dd03l_self:
            return jsonify(error='table_not_in_dd03l',
                           hint=f'No entries found in DD03L for table {table}. '
                                f'Upload the DD03L extract containing {table} first.'), 400
    else:
        threshold = 1.0 if table == 'DD03L' else 0.95
        non_empty_headers = [h for h in headers if h]
        matched_headers   = [h for h in non_empty_headers if h in tbl_fields]
        unmatched_headers = [(headers.index(h) + 1, h) for h in non_empty_headers if h not in tbl_fields]
        if not non_empty_headers or len(matched_headers) / len(non_empty_headers) < threshold:
            sample = [f'col {col} "{name}"' for col, name in unmatched_headers[:10]]
            return jsonify(error='non_technical_columns', table=table,
                           matched=len(matched_headers), total=len(non_empty_headers),
                           unmatched_sample=sample,
                           hint='Column headers must be SAP technical field names. '
                                'Re-export from SE16N using technical column names.'), 400

    enriched_columns = _enrich_columns(table, headers)

    # Dynamic table: DROP + CREATE + INSERT
    non_empty = [h for h in headers if h]
    col_defs  = ', '.join(f'"{h}" TEXT' for h in non_empty)
    col_list  = ', '.join(f'"{h}"' for h in non_empty)
    val_marks = ', '.join('?' * len(non_empty))

    c = harness_db()
    c.execute('BEGIN')
    c.execute(f'DROP TABLE IF EXISTS "{table}"')
    c.execute(f'CREATE TABLE "{table}" ({col_defs})')
    if rows:
        c.executemany(
            f'INSERT INTO "{table}" ({col_list}) VALUES ({val_marks})',
            [tuple(r.get(h, '') for h in non_empty) for r in rows]
        )
    c.execute('''
        INSERT OR REPLACE INTO _table_meta
            (tablename, system, client, date, project, filename, uploadedAt, columns, enriched_columns)
        VALUES (?,?,?,?,?,?,?,?,?)
    ''', (table, system, client, date, current_project, f.filename,
          datetime.utcnow().isoformat() + 'Z',
          json.dumps(headers), json.dumps(enriched_columns)))
    c.commit(); c.close()

    # Self-describing DD03L: also update _dd03l reference
    if table == 'DD03L' and not tbl_fields:
        norm = [{_DD03L_HEADER_MAP.get(k, k): v for k, v in r.items()} for r in rows]
        slim = []
        for r in norm:
            tn = str(r.get('TABNAME', '')).strip()
            fn = str(r.get('FIELDNAME', '')).strip()
            if not tn or not fn: continue
            slim.append({k: (str(v).strip() if v is not None else '') for k, v in r.items()})
        if slim:
            _write_dd03l(slim, f.filename)
            _reenrich_all()

    return jsonify(ok=True, table=table, rows=len(rows), columns=len(headers),
                   system=system, client=client, date=date)

# ───────────── DATA ─────────────
@app.get('/api/data/<table>')
@login_required
def get_data(table):
    table = table.upper()
    c = harness_db()
    meta = c.execute(
        'SELECT columns, enriched_columns, filename, uploadedAt, system, client, date '
        'FROM _table_meta WHERE tablename=?', (table,)
    ).fetchone()
    if not meta:
        c.close(); return jsonify(error='not_found'), 404
    columns  = json.loads(meta['columns'] or '[]')
    enriched = json.loads(meta['enriched_columns'] or '{}') or {}
    if not enriched:
        textfield = request.args.get('textfield', 'SCRTEXT_M')
        enriched  = _enrich_columns(table, columns, textfield)
    non_empty = [h for h in columns if h]
    try:
        col_sql  = ', '.join(f'"{h}"' for h in non_empty)
        rows_raw = c.execute(f'SELECT {col_sql} FROM "{table}"').fetchall()
    except sqlite3.OperationalError:
        c.close(); return jsonify(error='not_found'), 404
    c.close()
    rows = [{enriched.get(col, col): (row[i] if row[i] is not None else '')
             for i, col in enumerate(non_empty)}
            for row in rows_raw]
    cols    = [enriched.get(col, col) for col in columns]
    matched = sum(1 for col in columns if col in enriched)
    return jsonify(table=table, columns=cols, rows=rows,
                   matched=matched, total=len(columns),
                   filename=meta['filename'], uploadedAt=meta['uploadedAt'],
                   system=meta['system'], client=meta['client'], date=meta['date'])

@app.post('/api/data/<table>/describe')
@login_required
def describe_column(table):
    table = table.upper()
    d_req = request.get_json(silent=True) or {}
    field = str(d_req.get('field', '')).strip().upper()
    if not field: return jsonify(error='no_field'), 400

    c = harness_db()
    if not c.execute('SELECT 1 FROM _table_meta WHERE tablename=?', (table,)).fetchone():
        c.close(); return jsonify(error='not_found'), 404

    # Step 1: CHECKTABLE from _dd03l
    row = c.execute(
        'SELECT "CHECKTABLE" FROM _dd03l WHERE "TABNAME"=? AND "FIELDNAME"=? LIMIT 1',
        (table, field)
    ).fetchone()
    c.close()
    checktable = str(row[0]).strip() if row and row[0] else ''
    if not checktable:
        return jsonify(error='no_checktable',
                       hint=f'No check table defined for {table}.{field} in DD03L'), 400

    # Step 2: text table from _dd08l
    text_table = _dd08l_lookup(checktable)
    if not text_table:
        c = harness_db()
        has_dd08l = c.execute('SELECT 1 FROM _dd08l_meta WHERE k=?', ('uploadedAt',)).fetchone()
        c.close()
        if not has_dd08l:
            return jsonify(error='dd08l_missing',
                           hint='Upload DD08L to enable value description lookup'), 400
        return jsonify(error='no_text_table',
                       hint=f'No text table (FRKART=TEXT) found in DD08L for check table {checktable}'), 400

    # Step 3: load text table rows
    warning = None; values = {}
    c = harness_db()
    tt_meta = c.execute('SELECT columns FROM _table_meta WHERE tablename=?', (text_table,)).fetchone()
    if not tt_meta:
        warning = f'Upload {text_table} as a transactional table to get descriptions for {field}'
    else:
        tt_cols = json.loads(tt_meta['columns'] or '[]')
        has_spras = 'SPRAS' in tt_cols
        try:
            tt_rows = c.execute(f'SELECT * FROM "{text_table}"').fetchall()
            for row in tt_rows:
                row_dict = dict(zip(tt_cols, row))
                if has_spras:
                    lang = str(row_dict.get('SPRAS', '')).strip().upper()
                    if lang not in ('EN', 'E'): continue
                val   = str(row_dict.get(field, '')).strip()
                vtext = str(row_dict.get('VTEXT', '')).strip()
                if val and vtext:
                    values[val] = vtext
        except sqlite3.OperationalError:
            warning = f'Upload {text_table} as a transactional table to get descriptions for {field}'
    c.close()
    return jsonify(field=field, description_column=f'{field} - Description',
                   checktable=checktable, text_table=text_table,
                   values=values, warning=warning)

@app.delete('/api/data/<table>')
@login_required
def del_data(table):
    table = table.upper()
    c = harness_db()
    c.execute('BEGIN')
    c.execute(f'DROP TABLE IF EXISTS "{table}"')
    c.execute('DELETE FROM _table_meta WHERE tablename=?', (table,))
    c.commit(); c.close()
    return jsonify(ok=True)

# ───────────── ADMIN ─────────────
@app.post('/api/admin/migrate-legacy')
@login_required
def migrate_legacy():
    d = request.get_json(silent=True) or {}
    project = (d.get('project') or '').strip()
    if not project: return jsonify(error='missing_project'), 400
    c = harness_db()
    result = c.execute(
        'UPDATE _table_meta SET project=? WHERE project IS NULL OR project=""', (project,))
    c.commit(); c.close()
    return jsonify(ok=True, updated=result.rowcount)

# ───────────── STATIC ─────────────
@app.get('/')
def root():
    if not NO_AUTH and 'uid' not in session:
        return redirect('/login')
    if not NO_AUTH and 'project' not in session:
        return redirect('/login?step=project')
    return send_from_directory(HERE, 'index.html')

@app.get('/login')
def login_page():
    if NO_AUTH:
        return redirect('/')
    return send_from_directory(HERE, 'login.html')

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000, debug=True)
